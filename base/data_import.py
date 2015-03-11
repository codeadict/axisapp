import datetime
import cStringIO
import uuid
from dateutil.parser import parse
from celery import task

from tempfile import TemporaryFile
import openpyxl
from openpyxl.styles import Style, Border, Side, Alignment, Protection, Font
from openpyxl.cell import get_column_letter
from openpyxl.styles.numbers import NumberFormat
from openpyxl.worksheet.datavalidation import DataValidation

from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db.transaction import TransactionManagementError
from django.db import models, transaction, IntegrityError
from django.utils.translation import ugettext_lazy as _
from django.contrib import messages
from django.contrib.gis.geos import Point

from base import models as base_models
from censo.models import Cliente


IMPORT_CLIENT_FIELDS = [
    ('nombres', {'help_text': 'obligatorio'}),
    ('apellidos', {'help_text': 'obligatorio'}),
    ('tipo_id', {'column_name': 'Tipo Identificacion', 'help_text': 'obligatorio', 'choices': {'cedula': 0, 'ruc': 1, 'pasaporte': 2}}),
    ('identif', {'column_name': 'Identificacion', 'help_text': 'obligatorio'}),
    ('email', {'help_text': 'no es obligatorio'}),
    ('celular', {'help_text': 'no es obligatorio'}),
    ('convencional', {'help_text': 'no es obligatorio'}),
    ('cumple', {'help_text': 'no es obligatorio'}),
    ('administrador', {'help_text': 'si se deja en blanco toma nombres y apellidos'}),
    ('direccion', {'help_text': 'obligatorio'}),
    ('tipo_local', {'help_text': 'obligatorio', 'choices': {'arrendado': 0, 'propio': 1, 'comodato': 3}}),
    ('especial', {'help_text': 'si se deja en blanco sera falso'}),
    ('estatal', {'help_text': 'si se deja en blanco sera falso'}),
    ('persona_compras', {'help_text': 'si se deja en blanco toma nombres y apellidos'}),
    ('razon_social', {'help_text': 'si se deja en blanco toma nombres y apellidos'}),
    ('nombre_comercial', {'help_text': 'no es obligatorio'}),
    ('website', {'column_name': 'Sitio Web', 'help_text': 'no es obligatorio'}),
    ('macro_canal', {'help_text': 'no es obligatorio'}),
    ('ocasion_consumo', {'help_text': 'no es obligatorio'}),
    ('canal', {'help_text': 'no es obligatorio'}),
    ('subcanal', {'help_text': 'no es obligatorio'}),
    ('medida_frente', {'help_text': 'no es obligatorio'}),
    ('medida_fondo', {'help_text': 'no es obligatorio'}),
    ('horario_desde', {'help_text': 'obligatorio'}),
    ('horario_hasta', {'help_text': 'obligatorio'}),
    ('abc_compras', {'help_text': 'no es obligatorio', 'choices': {'a': 'a', 'b': 'b', 'c': 'c'}}),
    ('abc_industrias', {'help_text': 'no es obligatorio', 'choices': {'a': 'a', 'b': 'b', 'c': 'c'}}),
    ('codigo', {'help_text': 'no es obligatorio, se genera automaticamente'}),
    ('barrio', {'help_text': 'no es obligatorio'}),
    ('sector', {'help_text': 'no es obligatorio'}),
    ('coordenadas', {'help_text': 'no es obligatorio, usar latitud,longitud'}),
    ('estado', {'help_text': 'no es obligatorio, por defecto activo', 'choices': {'activo': 0, 'pasivo': 1, 'eliminado': 2}}),
]


TemporaryStorage = FileSystemStorage()

_import_cache_key = 'app-data-upload'


def _set_messages(request, result):
    for info in result['info']:
        messages.info(request, info)
    for warning in result['warnings']:
        messages.warning(request, warning)
    for e in result['errors']:
        messages.error(request, e)


def trigger_import(request, xl_file):
    if settings.USE_ASYNC_IMPORT:
        file_name = 'temporary-uploads/%s.xlsx' % (uuid.uuid4())
        TemporaryStorage.save(file_name, xl_file)
        job_id = _import_cache_key
        async_import.dalay(file_name)
    else:
        # this should work for tests and when running without worker.
        result = ImportData(xl_file).result()
        _set_messages(request, result)

@task()
def async_import(file_name):
    xl_file = TemporaryStorage.open(file_name)
    return ImportData(xl_file, True).result()


class ImportFormatError(Exception):
    pass


class ImportData(object):
    def __init__(self, xl_file, delete_after=False):
        self._info = []
        self._warnings = []
        self._errors = []
        clients_added, clients_updated = [], []
        self.current_row = None
        self.current_sheet = None
        try:
            with transaction.atomic():
                try:
                    # make sure file is local to speed up import when file was stored on S3
                    temp_file = TemporaryFile()
                    temp_file.write(xl_file.read())
                    temp_file.seek(0)
                    self.wb = openpyxl.load_workbook(temp_file, use_iterators=True, read_only=True, data_only=True)
                except AttributeError, e:
                    self.format_error(_(u'opening file "%(file)s", details: %(e)s') % {'file': xl_file, 'e': e},
                                      on_sheet=False)
                    raise ImportFormatError()
                clients_added, clients_updated = self.process_sheet(u'Clientes', Cliente, IMPORT_CLIENT_FIELDS)
                #if self.warnings():
                    #raise ImportFormatError()
        except ImportFormatError:
            self.errors(_(u'Error processing document, no records imported'))
        else:
            self.info(_(u'Import successful'))
            msg = _(u'%(added)d Clients added and %(updated)s updated')
            self.info(msg % {'added': len(clients_added), 'updated': len(clients_updated)})
        finally:
            if delete_after and hasattr(xl_file, 'name'):
                TemporaryStorage.delete(xl_file.name)

    def info(self, message=None):
        if message:
            self._info.append(message)
        return self._info

    def warnings(self, message=None):
        if message:
            self._warnings.append(message)
        return self._warnings

    def errors(self, message=None):
        if message:
            self._errors.append(message)
        return self._errors

    def result(self):
        return {
            'info': self.info(),
            'warnings': self.warnings(),
            'errors': self.errors()
        }

    def format_error(self, msg, field=None, on_sheet=True):
        if on_sheet:
            if field:
                prefix = _(u'sheet: %(sheet)s, row: %(row)s, column: %(column)s, error: %%s') % \
                    {'sheet': self.current_sheet, 'row': self.current_row, 'column': field}
            else:
                prefix = _(u'sheet: %(sheet)s, row: %(row)s, error: %%s') % \
                    {'sheet': self.current_sheet, 'row': self.current_row, 'column': field}
        else:
            prefix = _(u'error: %s')
        self.warnings(prefix % msg)

    def prepare_data(self, row, field_info, model_class):
        data = {}
        model_fields = {}
        for model_field in model_class._meta.fields:
            model_fields[model_field.name] = model_field

        for field, cell in zip(field_info, row):
            value = cell.value
            if not value:
                continue
            if isinstance(value, basestring):
                value = value.strip()
            field_name, field_info = field
            if field_name not in model_fields:
                self.format_error(_(u'Field "%s" not found in model fields.') % field_name)
                continue

            model_field = model_fields[field_name]

            if isinstance(model_field, models.ForeignKey) and model_field.rel.to == base_models.MacroCanal:
                try:
                    value = base_models.MacroCanal.objects.get(pk=value)
                except (base_models.MacroCanal.DoesNotExist, TransactionManagementError):
                    # i think TransactionManagementError is just a unit test thing but not clear
                    self.format_error(_(u'Associated macro channel "%s" not found') % value, field_name)
                    value = None
                except base_models.MacroCanal.MultipleObjectsReturned:
                    self.format_error(_(u'Multiple possible associated macro channels found for "%s"') % value)
                    return {}

            if isinstance(model_field, models.ForeignKey) and model_field.rel.to == base_models.OcasionConsumo:
                try:
                    value = base_models.OcasionConsumo.objects.get(pk=value)
                except (base_models.OcasionConsumo.DoesNotExist, TransactionManagementError):
                    # i think TransactionManagementError is just a unit test thing but not clear
                    self.format_error(_(u'Associated consumer ocassion "%s" not found') % value, field_name)
                    value = None
                except base_models.OcasionConsumo.MultipleObjectsReturned:
                    self.format_error(_(u'Multiple possible associated consumer ocassions found for "%s"') % value)
                    return {}

            if isinstance(model_field, models.ForeignKey) and model_field.rel.to == base_models.Canal:
                try:
                    value = base_models.Canal.objects.get(pk=value)
                except (base_models.Canal.DoesNotExist, TransactionManagementError):
                    # i think TransactionManagementError is just a unit test thing but not clear
                    self.format_error(_(u'Associated channel "%s" not found') % value, field_name)
                    value = None
                except base_models.Canal.MultipleObjectsReturned:
                    self.format_error(_(u'Multiple possible associated channels found for "%s"') % value)
                    return {}

            if isinstance(model_field, models.ForeignKey) and model_field.rel.to == base_models.SubCanal:
                try:
                    value = base_models.SubCanal.objects.get(pk=value)
                except (base_models.SubCanal.DoesNotExist, TransactionManagementError):
                    # i think TransactionManagementError is just a unit test thing but not clear
                    self.format_error(_(u'Associated sub-channel "%s" not found') % value, field_name)
                    continue
                except base_models.SubCanal.MultipleObjectsReturned:
                    self.format_error(_(u'Multiple possible associated sub-channels found for "%s"') % value)
                    return {}

            # Temporary fix for clients horario_desde y horario_hasta
            if isinstance(model_field, models.TimeField) and not value:
                value = datetime.time(0, 0, 0)

            if isinstance(model_field, models.DateField):
                try:
                    value = parse(unicode(value), yearfirst=True)
                except (TypeError, ValueError), e:
                    value = None
                    self.format_error(_(u'Invalid date: %s') % e, field_name)

            if isinstance(model_field, models.BooleanField) and not value:
                # TODO: Fix this to something better maybe
                value = False

            # For cordinate fields we use lat,lon and convert it to geodjango point
            if field_name in ['coordenadas', 'geom'] and value:
                lat, lon = value.split(',')
                try:
                    value = Point(float(lon), float(lat), srid=4326)
                except ValueError,e:
                    value = None

            choice_lookup = field_info.get('choices')
            if choice_lookup:
                real_value = choice_lookup.get(value.lower())
                if real_value is None:
                    keys = ', '.join(choice_lookup.keys())
                    self.format_error(_(u'Value %(value)s not found in choices, options: %(options)s') %
                                      {'value': value, 'options': keys}, field_name)
                    # prevent further errors on this, import will now fail anyway
                    real_value = choice_lookup[choice_lookup.keys()[0]]
                value = real_value
            else:
                if not model_field.blank and not value:
                    self.format_error(_(u'Field can\'t be empty'), field_name)
                else:
                    data[model_field.name] = value

        return data

    def process_sheet(self, sheet_name, model, fields):
        ws = self.wb.get_sheet_by_name(name=sheet_name)
        if ws is None:
            self.format_error(_(u'"%s" sheet not found.') % sheet_name, on_sheet=False)
            raise ImportFormatError()
        rows = ws.iter_rows()
        rows.next()  # Skipping headings
        rows.next()  # Skipping help text
        self.current_row = 2
        self.current_sheet = sheet_name
        clients_added = []
        clients_updated = []
        for row in rows:
            self.current_row += 1

            data = self.prepare_data(row, fields, model)
            if not data:
                continue
            print data

            # TODO: Uncomment this on future when add more models to import , now is giving transaction error, lets
            #       assume that the data is empty when importing

            # try:
            #     instance = model.objects.get(identif=data['identif'])
            # except model.DoesNotExist:
            #     try:
            instance = model.objects.create(**data)
                # except IntegrityError:
                #     self.format_error(_(u'Integrity Error, problem creating client.'))
                #     continue
            clients_added.append(instance.pk)
            # else:
            #     model.objects.filter(pk=instance.pk).update(**data)
            #     clients_updated.append(instance.pk)
        return clients_added, clients_updated


class TemplateSheet(object):
    @staticmethod
    def create(request):
        cls = TemplateSheet(request)
        return cls._get_doc()

    def __init__(self, request):
        self._wb = openpyxl.Workbook()
        [self._wb.remove_sheet(ws) for ws in self._wb.worksheets]
        self._add_sheet(u'Clientes', IMPORT_CLIENT_FIELDS, Cliente)

    def _get_doc(self):
        io = cStringIO.StringIO()
        self._wb.save(io)
        return io.getvalue()

    def _add_sheet(self, name, extra_fields, model_type):
        ws = self._wb.create_sheet()
        ws.title = name
        headings = self._get_headings(extra_fields, model_type)
        ws.append(headings)
        help_text = self._get_help_text(extra_fields)
        ws.append(help_text)
        self._set_sheet_styles(ws)
        for i, field in enumerate(extra_fields):
            col = get_column_letter(i + 1)
            _, extra = field
            if 'choices' in extra:
                choices = u','.join(c for c in extra['choices'])
                dv = DataValidation(type='list', formula1=u'"%s"' % choices, allow_blank=True)
                ws.add_data_validation(dv)
                dv.ranges.append(u'%s3:%s1000' % (col, col))

    @staticmethod
    def _get_headings(extra_fields, model_type):
        headings = []
        role_lookup = {f.name: f.verbose_name.encode('utf-8') for f in model_type._meta.fields}
        for field, extra in extra_fields:
            headings.append(extra.get('column_name') or role_lookup[field])
        return headings

    def _get_help_text(self, extra_fields):
        help_texts = []
        for _, extra in extra_fields:
            ht = extra.get('help_text', '')
            if not ht and 'choices' in extra:
                ht = u'Seleccione %s' % ', '.join(u'"%s"' % c for c in extra['choices'])
            help_texts.append(ht)
        return help_texts

    @staticmethod
    def _set_sheet_styles(ws):
        bottom_border = Border(bottom=Side(border_style='medium', color='FF000000'))
        locked = Protection(locked=True)
        row1 = ws.row_dimensions[1]
        row1.style = Style(font=Font(bold=True), protection=locked)
        row2 = ws.row_dimensions[2]
        row2.style = Style(font=Font(color='888888'), border=bottom_border, protection=locked,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        text_format = Style(number_format=NumberFormat.FORMAT_TEXT)
        for r in range(3, 1000):
            ws.append(('',))
            ws.row_dimensions[r].style = text_format
        row2.height = 60
        for col_dim in ws.column_dimensions.values():
            col_dim.width = 15
